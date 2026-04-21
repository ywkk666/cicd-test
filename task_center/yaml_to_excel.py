import os
import pandas as pd
from ruamel.yaml import YAML
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def sync_yaml_to_excel_pro(yaml_name, excel_name):
    """
    带全流程日志与异常捕获的同步脚本
    """
    # 1. 初始化路径
    base_path = os.path.dirname(os.path.abspath(__file__))
    yaml_path = os.path.join(base_path, yaml_name)
    excel_path = os.path.join(base_path, excel_name)

    print(f"🚀 开始同步流程...")
    print(f"📂 当前工作目录: {base_path}")

    # 2. 检查 YAML 文件
    if not os.path.exists(yaml_path):
        print(f"❌ 失败: 找不到文件 {yaml_name}。请确保它在 {base_path} 目录下。")
        return

    yaml = YAML()
    try:
        with open(yaml_path, 'r', encoding='utf-8') as f:
            data = yaml.load(f)
        issues_list = data.get('issues', [])
        if not issues_list:
            print("⚠️ 警告: YAML 文件中 'issues' 列表为空，没有数据可以同步。")
            return
        print(f"📖 成功读取 YAML，找到 {len(issues_list)} 个任务条目。")
    except Exception as e:
        print(f"❌ 解析 YAML 失败: {e}")
        return

    # 3. 读取 Excel
    if not os.path.exists(excel_path):
        print(f"❌ 失败: 目标 Excel 文件 {excel_name} 不存在。")
        return

    try:
        df_excel = pd.read_excel(excel_path)
        print(f"📊 成功加载 Excel，当前记录共 {len(df_excel)} 行。")

        # 4. 数据预处理
        df_yaml = pd.DataFrame(issues_list)
        
        # 清理 labels 的中括号
        if 'labels' in df_yaml.columns:
            df_yaml['labels'] = df_yaml['labels'].apply(lambda x: ', '.join(x) if isinstance(x, list) else x)

        # 检查关键列 'title' 是否存在 (作为对齐依据)
        if 'title' not in df_excel.columns:
            print(f"❌ 错误: Excel 中缺少关键列 'title'，无法进行数据匹配！")
            return

        # 5. 调整列顺序 & 类型对齐
        # 强制指定 ID 在首位，随后是 title 和 status
        fixed_cols = ['id', 'title', 'status']
        other_cols = [c for c in df_excel.columns if c not in fixed_cols]
        new_order = fixed_cols + other_cols

        for col in df_yaml.columns:
            if col in df_excel.columns:
                df_excel[col] = df_excel[col].astype(object)

        # 6. 执行同步更新
        df_excel.set_index('title', inplace=True)
        df_yaml.set_index('title', inplace=True)
        df_excel.update(df_yaml)
        df_excel.reset_index(inplace=True)

        # 应用新列序
        df_excel = df_excel[new_order]

        # 保存基础数据 (这一步必须先成功，后面才能上色)
        df_excel.to_excel(excel_path, index=False)
        print(f"💾 基础数据已回写至 Excel，列顺序已重置（ID 居首）。")

        # 7. 渲染样式 (上色)
        print(f"🎨 正在进行状态上色处理...")
        wb = load_workbook(excel_path)
        ws = wb.active

        # 定义颜色
        colors = {
            'done': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),       # 绿
            'processing': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"), # 黄
            'error': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")       # 红
        }

        # 定位 status 列索引
        status_idx = None
        for cell in ws[1]:
            if cell.value == 'status':
                status_idx = cell.column
                break

        if status_idx:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=status_idx)
                status_val = str(cell.value).lower().strip() if cell.value else ""
                
                if status_val == 'done':
                    cell.fill = colors['done']
                elif status_val == 'processing':
                    cell.fill = colors['processing']
                else:
                    cell.fill = colors['error']
        
        wb.save(excel_path)
        print(f"✨ 全流程圆满完成！请打开 {excel_name} 查看结果。")

    except PermissionError:
        print(f"⚠️ 同步失败: 权限拒绝！请检查是否打开了 {excel_name}，请关闭它后重试。")
    except Exception as e:
        print(f"❌ 运行中发生未预期错误: {e}")

if __name__ == "__main__":
    sync_yaml_to_excel_pro('tasks.yaml', 'task_manager.xlsx')